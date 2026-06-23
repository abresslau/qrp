"""BoE workbook parser — pure, no network. Builds synthetic workbooks matching BoE's layout."""

from __future__ import annotations

import io
import zipfile
from datetime import datetime

import openpyxl
import pytest

from rates.sources.boe import CurveLayoutError, _parse_zip_bytes, parse_workbook

# (sheet title, has 'years:' header) — BoE's four curve sheets.
_SHEETS = ["1. fwds, short end", "2. fwd curve", "3. spot, short end", "4. spot curve"]


def _wb(sheet_specs: dict) -> io.BytesIO:
    """sheet_specs: {sheet_name: (tenors, [(date, [values...]), ...])}. Adds an 'info' sheet."""
    wb = openpyxl.Workbook()
    wb.active.title = "info"
    wb["info"]["A1"] = "Bank of England UK yield curve data"
    for name, (tenors, rows) in sheet_specs.items():
        ws = wb.create_sheet(name)
        ws["A1"] = "title"
        ws["A4"] = "years:"  # the tenor header row (col A label, tenors from col B)
        for j, t in enumerate(tenors):
            ws.cell(row=4, column=2 + j, value=t)
        for i, (d, vals) in enumerate(rows):
            ws.cell(row=6 + i, column=1, value=d)
            for j, v in enumerate(vals):
                if v is not None:
                    ws.cell(row=6 + i, column=2 + j, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_parses_spot_and_forward_with_dates_and_tenors():
    d = datetime(2026, 6, 1)
    buf = _wb({
        "4. spot curve": ([0.5, 1.0], [(d, [3.5, 3.7])]),
        "2. fwd curve": ([0.5, 1.0], [(d, [3.4, 3.9])]),
    })
    pts = parse_workbook(buf, "glc", "nominal")
    spot = {(p.rate_type, p.tenor): p.value for p in pts}
    assert spot[("spot", 0.5)] == 3.5
    assert spot[("spot", 1.0)] == 3.7
    assert spot[("forward", 0.5)] == 3.4
    assert all(p.curve_set == "glc" and p.basis == "nominal" for p in pts)
    assert all(p.as_of_date == d.date() for p in pts)


def test_short_end_dedupes_against_curve_grid():
    d = datetime(2026, 6, 1)
    # curve grid has 0.5, 1.0; short-end has 0.0833 (new) + 0.5 (overlap → curve wins)
    buf = _wb({
        "4. spot curve": ([0.5, 1.0], [(d, [3.5, 3.7])]),
        "3. spot, short end": ([0.083333, 0.5], [(d, [3.0, 9.99])]),  # 9.99 must NOT win at 0.5
    })
    pts = [p for p in parse_workbook(buf, "glc", "nominal") if p.rate_type == "spot"]
    by_t = {p.tenor: p.value for p in pts}
    assert by_t[0.5] == 3.5  # curve grid wins over the short-end's 9.99
    assert by_t[1.0] == 3.7
    assert by_t[round(0.083333, 6)] == 3.0  # short-end-only sub-year node added
    assert len(pts) == 3


def test_skips_non_numeric_cells():
    d = datetime(2026, 6, 2)
    buf = _wb({"4. spot curve": ([0.5, 1.0, 1.5], [(d, [3.5, None, 3.9])])})
    pts = parse_workbook(buf, "ois", "nominal")
    tenors = sorted(p.tenor for p in pts)
    assert tenors == [0.5, 1.5]  # the None (missing) tenor is dropped, not zero-filled


def test_missing_years_header_raises_layout_error():
    wb = openpyxl.Workbook()
    wb.active.title = "info"
    ws = wb.create_sheet("4. spot curve")
    ws["A1"] = "title"  # no 'years:' row anywhere
    ws["A6"] = datetime(2026, 6, 1)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with pytest.raises(CurveLayoutError):
        parse_workbook(buf, "glc", "nominal")


def test_no_expected_sheets_raises_layout_error():
    wb = openpyxl.Workbook()
    wb.active.title = "info"
    wb.create_sheet("something else")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with pytest.raises(CurveLayoutError):
        parse_workbook(buf, "glc", "nominal")


def test_parse_zip_recurses_into_nested_zip():
    """Regression: the BoE 'latest' bundle wraps the four xlsx inside an INNER zip (plus .gif
    previews). A flat top-level scan parses 0 points and the daily load silently no-ops — so the
    parser must recurse into the nested zip and ignore the gifs."""
    d = datetime(2026, 6, 22)
    xlsx = _wb({
        "4. spot curve": ([1.0, 10.0], [(d, [4.0, 4.84])]),
        "2. fwd curve": ([1.0, 10.0], [(d, [4.1, 4.9])]),
    }).getvalue()
    inner = io.BytesIO()
    with zipfile.ZipFile(inner, "w") as z:
        z.writestr("GLC Nominal daily data current month.xlsx", xlsx)
    outer = io.BytesIO()
    with zipfile.ZipFile(outer, "w") as z:
        z.writestr("Latest Yield Curve data (current month).zip", inner.getvalue())
        z.writestr("uknom.gif", b"not a curve")  # must be ignored, not parsed

    pts = _parse_zip_bytes(outer.getvalue())
    assert pts, "nested-zip xlsx must be parsed (BoE wraps the daily xlsx in an inner zip)"
    assert all(p.curve_set == "glc" and p.basis == "nominal" for p in pts)
    assert {p.as_of_date for p in pts} == {d.date()}
