"""Bank of England UK yield-curve source adapter.

BoE's Monetary & Financial Conditions Division publishes daily UK yield curves as Excel files
bundled in zips (Open Government Licence — free, attribution). Probed 2026-06-22:

  - ``latest-yield-curve-data.zip`` — the CURRENT MONTH daily files (4 xlsx: GLC Nominal/Real/
    Inflation + OIS).
  - ``glcnominalddata.zip`` / ``glcrealddata.zip`` / ``glcinflationddata.zip`` / ``oisddata.zip``
    — the FULL daily history per curve (large; ~39 MB for nominal).

Each xlsx has sheets ``1. fwds, short end`` / ``2. fwd curve`` / ``3. spot, short end`` /
``4. spot curve``. A sheet's row 'years:' carries the tenor grid (years); subsequent rows are
``<date>, v(t0), v(t1), …`` with values in **% per annum** (real can be negative). The stated
date in column A is the curve's ``as_of_date`` (never the ingest date).

This module separates **parsing** (pure, testable, no network) from **downloading**.
"""

from __future__ import annotations

import io
import zipfile
from datetime import date, datetime
from pathlib import Path
from urllib.request import Request, urlopen

import openpyxl

from .base import CurvePoint

BOE_BASE = "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves"
LATEST_ZIP = f"{BOE_BASE}/latest-yield-curve-data.zip"

# Full daily-history archives, keyed by (curve_set, basis).
ARCHIVE_ZIPS: dict[tuple[str, str], str] = {
    ("glc", "nominal"): f"{BOE_BASE}/glcnominalddata.zip",
    ("glc", "real"): f"{BOE_BASE}/glcrealddata.zip",
    ("glc", "inflation"): f"{BOE_BASE}/glcinflationddata.zip",
    ("ois", "nominal"): f"{BOE_BASE}/oisddata.zip",
}

# Map a workbook filename (substring match, case-insensitive) → (curve_set, basis).
FILE_MAP: list[tuple[str, tuple[str, str]]] = [
    ("glc inflation", ("glc", "inflation")),
    ("glc nominal", ("glc", "nominal")),
    ("glc real", ("glc", "real")),
    ("ois", ("ois", "nominal")),
]

# Sheet name (substring, lowercased) → (rate_type, is_curve_grid). The "curve" grids are the
# canonical full-span curves (0.5y steps); the "short end" grids add finer sub-year resolution.
SHEET_MAP: list[tuple[str, tuple[str, bool]]] = [
    ("fwd curve", ("forward", True)),
    ("spot curve", ("spot", True)),
    ("fwds, short end", ("forward", False)),
    ("spot, short end", ("spot", False)),
]

_UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) QRP-rates/0.1"


class CurveLayoutError(RuntimeError):
    """BoE's file layout drifted from what the probe recorded (fail loud, never mis-map)."""


def _curve_set_basis(filename: str) -> tuple[str, str] | None:
    low = filename.lower()
    for needle, cb in FILE_MAP:
        if needle in low:
            return cb
    return None


def _parse_sheet(ws, curve_set: str, basis: str, rate_type: str) -> list[CurvePoint]:
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(
        (
            i
            for i, r in enumerate(rows[:10])
            if r and isinstance(r[0], str) and "years" in r[0].lower()
        ),
        None,
    )
    if header_idx is None:
        raise CurveLayoutError(
            f"{curve_set}/{basis}/{rate_type}: no 'years:' tenor header in the first 10 rows"
        )
    header = rows[header_idx]
    # tenor columns: positions after col A that carry a numeric tenor
    tenor_cols = [
        (j, float(v)) for j, v in enumerate(header) if j >= 1 and isinstance(v, (int, float))
    ]
    if not tenor_cols:
        raise CurveLayoutError(f"{curve_set}/{basis}/{rate_type}: no numeric tenors in header")
    out: list[CurvePoint] = []
    for r in rows[header_idx + 1 :]:
        if not r or not isinstance(r[0], (datetime, date)):
            continue
        d = r[0].date() if isinstance(r[0], datetime) else r[0]
        for j, tenor in tenor_cols:
            if j >= len(r):
                continue
            v = r[j]
            if isinstance(v, (int, float)):
                out.append(
                    CurvePoint(
                        "GB", "GBP", curve_set, basis, rate_type, round(tenor, 6), d, float(v)
                    )
                )
    return out


def parse_workbook(path: str | Path, curve_set: str, basis: str) -> list[CurvePoint]:
    """Parse one BoE xlsx into curve points. Pure (no network). Asserts the expected sheets.

    Both the "curve" grid and the "short end" grid are kept, deduped so a tenor present in the
    canonical curve grid wins over the same tenor in the short-end grid (one curve, no PK clash);
    short-end-only tenors (the finer sub-year nodes) are added.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheets_low = {s.lower(): s for s in wb.sheetnames}

        curve_pts: list[CurvePoint] = []
        short_pts: list[CurvePoint] = []
        matched = 0
        for needle, (rate_type, is_curve) in SHEET_MAP:
            sheet = next((orig for low, orig in sheets_low.items() if needle in low), None)
            if sheet is None:
                continue
            matched += 1
            pts = _parse_sheet(wb[sheet], curve_set, basis, rate_type)
            (curve_pts if is_curve else short_pts).append(pts)
        if matched == 0:
            raise CurveLayoutError(
                f"{curve_set}/{basis}: none of the expected sheets {[n for n, _ in SHEET_MAP]} "
                f"found in {list(wb.sheetnames)}"
            )
    finally:
        wb.close()  # read-only workbooks hold the underlying zip handle until closed

    flat_curve = [p for grp in curve_pts for p in grp]
    seen = {(p.rate_type, p.as_of_date, p.tenor) for p in flat_curve}
    extra = [
        p for grp in short_pts for p in grp if (p.rate_type, p.as_of_date, p.tenor) not in seen
    ]
    return flat_curve + extra


def _parse_zip_bytes(blob: bytes, _depth: int = 0) -> list[CurvePoint]:
    """Parse the BoE curve xlsx out of a zip — recursing into NESTED zips.

    The ``latest-yield-curve-data.zip`` bundle wraps the four daily xlsx inside an inner
    ``Latest Yield Curve data (current month).zip`` (alongside some .gif previews), so a flat
    top-level scan finds no xlsx and silently parses 0 points. We recurse one level (bounded) so
    both the nested 'latest' bundle and the flat per-curve archive zips work."""
    if _depth > 3:  # bounded guard against a pathological nested zip; BoE nests exactly one level
        return []
    out: list[CurvePoint] = []
    with zipfile.ZipFile(io.BytesIO(blob)) as zf:
        for name in zf.namelist():
            low = name.lower()
            if low.endswith(".zip"):
                with zf.open(name) as fh:
                    out.extend(_parse_zip_bytes(fh.read(), _depth + 1))
                continue
            if not low.endswith((".xlsx", ".xls")):
                continue
            cb = _curve_set_basis(Path(name).name)
            if cb is None:
                continue
            with zf.open(name) as fh:
                out.extend(parse_workbook(io.BytesIO(fh.read()), cb[0], cb[1]))
    return out


def _download(url: str, *, timeout: int = 120) -> bytes:
    req = Request(url, headers={"User-Agent": _UA})
    with urlopen(req, timeout=timeout) as resp:  # noqa: S310 (trusted BoE host)
        return resp.read()


class BoeCurveSource:
    """Fetches + parses BoE UK yield curves. ``SOURCE`` tags every stored row."""

    SOURCE = "boe"
    COUNTRY = "GB"
    CURRENCY = "GBP"

    def __init__(self, *, archive: bool = False) -> None:
        # archive=False → the latest (current-month) bundle (the daily/tail case);
        # archive=True  → the full per-curve history zips (the backfill case).
        self.archive = archive

    def fetch(
        self, *, start_date: date | None = None, end_date: date | None = None
    ) -> list[CurvePoint]:
        pts: list[CurvePoint]
        if self.archive:
            pts = []
            for url in ARCHIVE_ZIPS.values():
                pts.extend(_parse_zip_bytes(_download(url)))
        else:
            pts = _parse_zip_bytes(_download(LATEST_ZIP))
        if start_date is not None:
            pts = [p for p in pts if p.as_of_date >= start_date]
        if end_date is not None:
            pts = [p for p in pts if p.as_of_date <= end_date]
        return pts
