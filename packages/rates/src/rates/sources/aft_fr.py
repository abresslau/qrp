"""Agence France Trésor (FR) inflation-linked (OAT€i) real-yield + breakeven source adapter.

France's nominal government curve is no longer published as a free daily multi-tenor feed (Banque de
France discontinued OAT rates 2024-07-10; AFT's full-grid workbook is a frozen 2021 demo). But AFT
DOES publish, daily, the euro-area **10-year benchmark** inflation-linked series as an .xlsx (served
with a legacy ``.xls`` extension). Probed 2026-07-01:

  ``https://www.aft.gouv.fr/files/medias-aft/3_Dette/3.3_OATEi/{YYYY}_{MM}_01_rend_tit_ref_oatei.xls``

The filename encodes the PUBLICATION month; each file carries the full daily history (2013-11 →) up
to its publish date. We resolve the latest available month by walking back from today (the current
month often 403s until published). Sheet "Données", after a title/header preamble, is one row per
business day:

  col 0 = date, col 1 = OAT€i **real** yield (10y), col 2 = nominal OAT yield (10y),
  col 3 = **breakeven** inflation (10y) — and by construction col1 + col3 = col2.

Values are **decimal fractions** (0.01462 = 1.462% p.a.), so we scale ×100 to the store's %-p.a.
convention. We emit the REAL yield (``basis='real'``) and the published BREAKEVEN
(``basis='inflation'``) at a nominal 10y tenor — AFT rolls the reference bond to keep it ~10y, so it
is a constant-~10y benchmark, not a fitted curve. The nominal column is NOT emitted (FR nominal
stays on the ECB series, and storing it would collide with the ECB Maastricht 10y point).

HONESTY: the OAT€i is linked to **euro-area HICP** (not French CPI), so the breakeven is EU
inflation; and this is a single benchmark point, not a fitted real term structure. This module
separates **parsing** (pure, testable, no network) from **downloading**.
"""

from __future__ import annotations

import io
import warnings
from datetime import date
from urllib.error import HTTPError
from urllib.request import Request, urlopen

import openpyxl

from .base import CurvePoint

_AFT_URL_TMPL = (
    "https://www.aft.gouv.fr/files/medias-aft/3_Dette/3.3_OATEi/"
    "{y:04d}_{m:02d}_01_rend_tit_ref_oatei.xls"
)
_TENOR = 10.0  # the "zone euro 10 ans" benchmark (AFT rolls the reference OAT€i to stay ~10y)
_UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) QRP-rates/0.1"


class CurveLayoutError(RuntimeError):
    """The AFT workbook layout drifted from what the probe recorded (fail loud, never mis-map)."""


def parse_workbook(source: str | io.BytesIO) -> list[CurvePoint]:
    """Parse the AFT OAT€i workbook into FR real + breakeven (inflation) 10y points. No network.

    Data rows are those whose first cell is a date; col1 = real yield, col3 = breakeven, both
    decimal fractions scaled x100. A blank cell is skipped (early history has breakeven, no real).
    """
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    def _num(cell) -> float | None:
        return float(cell) if isinstance(cell, (int, float)) else None

    def _as_date(cell) -> date | None:
        if isinstance(cell, date):
            return cell if not hasattr(cell, "date") else cell.date()
        return None

    out: list[CurvePoint] = []
    seen_date = False
    for r in rows:
        if not r:
            continue
        d = _as_date(r[0])
        if d is None:
            continue
        seen_date = True
        real = _num(r[1]) if len(r) > 1 else None
        be = _num(r[3]) if len(r) > 3 else None
        if real is not None:
            out.append(CurvePoint("FR", "EUR", "govt", "real", "yield", _TENOR, d, real * 100.0))
        if be is not None:
            out.append(
                CurvePoint("FR", "EUR", "govt", "inflation", "yield", _TENOR, d, be * 100.0))
    if not seen_date:
        raise CurveLayoutError("no date rows found in the AFT OAT€i workbook")
    return out


def _download_latest(*, today: date, timeout: int = 120, max_back: int = 4) -> bytes:
    """Fetch the newest available monthly file, walking back from ``today`` (the current month often
    404/403s until published). Raises if none of the last ``max_back`` months is reachable."""
    last_err: Exception | None = None
    y, m = today.year, today.month
    for _ in range(max_back):
        url = _AFT_URL_TMPL.format(y=y, m=m)
        try:
            req = Request(url, headers={"User-Agent": _UA})
            with urlopen(req, timeout=timeout) as resp:  # noqa: S310 (trusted AFT host)
                data = resp.read()
            if data[:2] == b"PK":  # a real .xlsx zip (not an HTML 403/404 page)
                return data
            last_err = CurveLayoutError(f"{url} returned a non-xlsx body")
        except (HTTPError, OSError) as exc:
            last_err = exc
        m -= 1
        if m == 0:
            y, m = y - 1, 12
    raise last_err or CurveLayoutError("no AFT OAT€i file reachable in the window")


class AftOateiCurveSource:
    """Fetches + parses the AFT OAT€i 10y real yield + published breakeven. ``SOURCE`` tags rows."""

    SOURCE = "aft"
    COUNTRY = "FR"
    CURRENCY = "EUR"

    def fetch(
        self, *, start_date: date | None = None, end_date: date | None = None
    ) -> list[CurvePoint]:
        anchor = end_date or date.today()
        pts = parse_workbook(io.BytesIO(_download_latest(today=anchor)))
        if start_date is not None:
            pts = [p for p in pts if p.as_of_date >= start_date]
        if end_date is not None:
            pts = [p for p in pts if p.as_of_date <= end_date]
        return pts
