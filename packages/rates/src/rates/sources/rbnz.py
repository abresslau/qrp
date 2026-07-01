"""Reserve Bank of New Zealand Government-bond-yield source adapter (statistical table B2).

The RBNZ publishes table B2 "Wholesale interest rates" (daily close) as an .xlsx that requires a
browser User-Agent. Probed 2026-06-22 (200, ~400 KB):

  ``https://www.rbnz.govt.nz/-/media/project/sites/rbnz/files/statistics/series/b/b2/hb2-daily-close.xlsx``

The "Data" sheet header is four rows: row 0 = group (e.g. "Secondary market government bond
closing"), row 1 = the per-column label ("1 year"/"2 year"/… or a specific bond maturity like
"May 2030"), row 4 = the ``Series Id`` (``INM.DG10<n>.NZZCF`` are the constant-maturity nominal
benchmarks). Data rows (date in col A) start at row 5. We keep the constant-maturity nominal
government benchmarks (1/2/5/10 year) PLUS the **inflation-indexed (real) government bonds**
(``INM.DG29.NS*ZC`` — dated linkers maturing Sep-2030/2035/2040/2050, emitted with ``basis='real'``
and a tenor computed per day from the bond's maturity date). Bank-bill yields and swap rates are
excluded.

NB: this workbook ships without a stored worksheet dimension, so it must be opened WITHOUT
``read_only`` (else openpyxl reports a single phantom row). This module separates **parsing**
(pure, no network) from **downloading**.
"""

from __future__ import annotations

import io
import re
import warnings
from datetime import date, datetime
from pathlib import Path
from urllib.request import Request, urlopen

import openpyxl

from .base import CurvePoint

RBNZ_URL = (
    "https://www.rbnz.govt.nz/-/media/project/sites/rbnz/files/"
    "statistics/series/b/b2/hb2-daily-close.xlsx"
)

# Constant-maturity nominal GoNZ benchmark Series Id → tenor in years (DG101/DG102/DG105/DG110).
SERIES_TENORS: dict[str, float] = {
    "INM.DG101.NZZCF": 1.0,
    "INM.DG102.NZZCF": 2.0,
    "INM.DG105.NZZCF": 5.0,
    "INM.DG110.NZZCF": 10.0,
}

# Inflation-indexed (real) GoNZ bonds — dated linkers (NOT constant-maturity), Series Id → maturity
# date. Tenor is computed per observation as (maturity − as_of_date)/365.25; emitted basis='real'.
# (The Sep-2025 linker INM.DG29.NS2509ZC has matured — its column is now empty and contributes
# nothing; new linkers can be added here as they're issued.)
REAL_SERIES_MATURITY: dict[str, date] = {
    "INM.DG29.NS3009ZC": date(2030, 9, 20),
    "INM.DG29.NS3509ZC": date(2035, 9, 20),
    "INM.DG29.NS4009ZC": date(2040, 9, 20),
    "INM.DG29.NS5009ZCF": date(2050, 9, 20),
}

# RBNZ's WAF rejects a non-browser User-Agent (a custom suffix → 403). Use a clean browser string.
_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)
_DATA_SHEET = "Data"
_SID_LABEL = "Series Id"


class CurveLayoutError(RuntimeError):
    """RBNZ's B2 layout drifted from what the probe recorded (fail loud, never mis-map)."""


def _coerce_date(cell) -> date | None:
    if isinstance(cell, datetime):
        return cell.date()
    if isinstance(cell, date):
        return cell
    if isinstance(cell, str):
        m = re.match(r"^(\d{4}-\d{2}-\d{2})", cell.strip())
        if m:
            return date.fromisoformat(m.group(1))
    return None


def parse_workbook(source: str | Path | io.BytesIO) -> list[CurvePoint]:
    """Parse a B2 workbook into nominal GoNZ benchmark yield points. Pure (no network)."""
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")  # "no default style" on RBNZ's bare workbook
        wb = openpyxl.load_workbook(source, data_only=True)  # not read_only: dimension is absent
    try:
        if _DATA_SHEET not in wb.sheetnames:
            raise CurveLayoutError(f"no '{_DATA_SHEET}' sheet in {wb.sheetnames}")
        rows = list(wb[_DATA_SHEET].iter_rows(values_only=True))
    finally:
        wb.close()
    sid_row_idx = next(
        (
            i
            for i, r in enumerate(rows[:10])
            if r and isinstance(r[0], str) and r[0].strip() == _SID_LABEL
        ),
        None,
    )
    if sid_row_idx is None:
        raise CurveLayoutError(f"no '{_SID_LABEL}' header row in the first 10 rows")
    sid_row = rows[sid_row_idx]
    col_tenors = [
        (j, SERIES_TENORS[str(sid).strip()])
        for j, sid in enumerate(sid_row)
        if j >= 1 and isinstance(sid, str) and sid.strip() in SERIES_TENORS
    ]
    # real (inflation-indexed) columns: column index → the bond's maturity date.
    col_real = [
        (j, REAL_SERIES_MATURITY[str(sid).strip()])
        for j, sid in enumerate(sid_row)
        if j >= 1 and isinstance(sid, str) and sid.strip() in REAL_SERIES_MATURITY
    ]
    if not col_tenors and not col_real:
        raise CurveLayoutError(f"no known GoNZ benchmark Series Ids in {sid_row}")
    out: list[CurvePoint] = []
    for r in rows[sid_row_idx + 1 :]:
        if not r:
            continue
        d = _coerce_date(r[0])
        if d is None:
            continue
        for j, tenor in col_tenors:
            if j >= len(r):
                continue
            v = r[j]
            if not isinstance(v, (int, float)):
                continue
            out.append(
                CurvePoint("NZ", "NZD", "govt", "nominal", "yield", tenor, d, float(v))
            )
        for j, maturity in col_real:
            if j >= len(r):
                continue
            v = r[j]
            if not isinstance(v, (int, float)):
                continue
            tenor = (maturity - d).days / 365.25
            if tenor <= 0:  # bond matured — skip
                continue
            out.append(
                CurvePoint("NZ", "NZD", "govt", "real", "yield", round(tenor, 6), d, float(v))
            )
    return out


def _download(url: str, *, timeout: int = 120) -> bytes:
    # RBNZ sits behind a WAF that 403s anything that doesn't look like a real browser navigation —
    # a full browser header set (Accept / Accept-Language / Referer) gets through where a bare UA
    # does not.
    req = Request(
        url,
        headers={
            "User-Agent": _UA,
            "Accept": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
                "application/octet-stream,*/*"
            ),
            "Accept-Language": "en-NZ,en;q=0.9",
            "Referer": "https://www.rbnz.govt.nz/statistics/series/exchange-and-interest-rates/"
            "wholesale-interest-rates",
        },
    )
    with urlopen(req, timeout=timeout) as resp:  # noqa: S310 (trusted RBNZ host)
        return resp.read()


class RbnzCurveSource:
    """Fetches + parses RBNZ B2 New Zealand Government yields. ``SOURCE`` tags every stored row."""

    SOURCE = "rbnz"
    COUNTRY = "NZ"
    CURRENCY = "NZD"

    def fetch(
        self, *, start_date: date | None = None, end_date: date | None = None
    ) -> list[CurvePoint]:
        pts = parse_workbook(io.BytesIO(_download(RBNZ_URL)))
        if start_date is not None:
            pts = [p for p in pts if p.as_of_date >= start_date]
        if end_date is not None:
            pts = [p for p in pts if p.as_of_date <= end_date]
        return pts
